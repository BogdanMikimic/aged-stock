class SalesPeople():
    #returneaza o lista de dictionare
    def __init__(self, file_with_path):
        from openpyxl import load_workbook
        global load_workbook
        self.unchecked_path = file_with_path
        self.my_xcel = None
        self.excel_object = None
        self.work_tab = None
        self.table_headings_needed = ['Customer Name', 'Customer Number', 'Sales Rep', 'Customer Care Agent']
        self.lista_dictionare = list()
        self.dictionar = dict()
        self.row_count = 0
        self.empty_rows_to_be_deleted = list()
        self.is_row_empty = False
        self.empty_cols_to_be_deleted = list()
        self.is_column_empty = False
        self.col_count = 0
        self.heading_not_found = list()
        self.my_xcel = self.unchecked_path


    def checkPosibleMultipleTabs(self):
        '''Checks file to have only one tab.
        Is not designed to work with more than one tab'''
        self.excel_object = load_workbook(self.my_xcel)
        if len(self.excel_object.sheetnames) == 1:
            self.work_tab = self.excel_object.active #get active tab
        else:
            return False

    def checkAndEliminateEmptyRows(self):
        '''Checks file and eliminates empty rows'''
        for row in self.work_tab.iter_rows():
            self.row_count += 1
            for itm in row:
                if itm.value != None:
                    self.is_row_empty = False
                    break
                else:
                    self.is_row_empty = True

            if self.is_row_empty == True:
                self.empty_rows_to_be_deleted.append(self.row_count)
                self.is_row_empty == False

        self.row_count = 0

        #delete empty rows
        if len(self.empty_rows_to_be_deleted) > 0:
            for empty_row in self.empty_rows_to_be_deleted:
                self.work_tab.delete_rows(empty_row)

    def checkAndEliminateEmptyColumns(self):
        '''Checks file and eliminates empty columns'''
        for col in self.work_tab.iter_cols():
            self.col_count += 1
            for itm1 in col:
                if itm1.value != None:
                    self.is_column_empty = False
                    break
                else:
                    self.is_column_empty = True

            if self.is_column_empty == True:
                self.empty_cols_to_be_deleted.append(self.col_count)
                self.is_column_empty == False

        self.col_count = 0

        #delete empty columns
        if len(self.empty_cols_to_be_deleted) > 0:
            for col_to_be_deleted in self.empty_cols_to_be_deleted:
                self.work_tab.delete_cols(col_to_be_deleted)

    def checkThatCorrectTabsAreContained(self):
        pass

        '''Check that the correct tabs are there'''
        self.row_count = 0
        for line in self.work_tab:
            self.row_count += 1
            if self.row_count == 1:
                for itm in line:
                    if itm not in self.table_headings_needed:
                        self.heading_not_found.append(itm)
            else:
                self.row_count = 0
                break
        if len(self.heading_not_found) > 0:
            return f'The program expects to find the following headings in the spreadsheet: Customer Name, Customer Number, Sales Rep, Customer Care Agent. They are case sensitive and there should not be a blank space before or after each heading'

    def returnTableAsDict(self):
        #need to check that each heading is named as it should
        self.row_count = 0
        for row in self.work_tab.iter_rows():
            self.row_count += 1
            if self.row_count > 1:
                itm_count = 0
                for itmn in row:
                    self.dictionar[self.table_headings_needed[itm_count]] = itmn.value
                    itm_count += 1
                self.lista_dictionare.append(self.dictionar)
                self.dictionar = dict()
        return self.lista_dictionare

    def runAll(self):
        if self.checkPosibleMultipleTabs() == False:
            return 'File has more than one tab. It only knows how to work if the data is in only one tab - where all the data is'
        self.checkAndEliminateEmptyRows()
        self.checkAndEliminateEmptyColumns()
        return self.returnTableAsDict()
