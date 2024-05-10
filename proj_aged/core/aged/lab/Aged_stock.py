class AgedStock():
    # ia ca argument fisierul excel cu aged stock, o citeste, si o intoarce ca lista de dictionare
    # prin optiunea load_workbook(fisier.xlsx, data_only=True) -> data_only=True aduce valorile in loc de formule
    # ultima coloana e vlookup (deci formula) si aduce valorile
    def __init__(self, file_with_path):
        from openpyxl import load_workbook
        import datetime
        global load_workbook, datetime
        self.my_xcel = file_with_path
        self.excel_object = None
        self.my_tab = None
        self.table_as_list_of_dicts = list()
        self.row_count = 0
        self.cell_count = 0


    def checkFileHasCorrectTabs(self):
        #expectation: table has tab 1 called pivot, tab 2 called data -> wee need data
        pass

    def checkIfTabHasTheRight(self):
        #expectation: headings are called 'Plnt' 'Stor loc' 'Material' 'Brand' 'Matl Group' 'Batch' 'Quantity' 'Unrestricted' 'Expiration date' 'Description'
        pass


    def runAll(self):
        #Create excel object
        self.excel_object = load_workbook(self.my_xcel, data_only=True)
        #Select right tab
        self.my_tab = self.excel_object['data']

        #itereaza rand cu rand
        for row in self.my_tab.iter_rows():
            self.row_count += 1
            if self.row_count >1: #treci peste primul rand care e data
                dictionar_row = dict() #creaza un dict care se refreshuie cu fiecare rand
                self.cell_count = 0
                #pune in dictionar datele din fiecare rand si fiecare dictionar in lista
                for cell in row:
                    self.cell_count += 1
                    if self.cell_count == 1:
                        dictionar_row['Plnt'] = cell.value
                    elif self.cell_count == 2:
                        dictionar_row['Stor loc'] = cell.value
                    elif self.cell_count == 3:
                        dictionar_row['Material'] = cell.value
                    elif self.cell_count == 4:
                        dictionar_row['Brand'] = cell.value
                    elif self.cell_count == 5:
                        # I have a material group called MISCEl. with a dot.
                        # That shits on my logic within using the material name
                        # as an HTML custom atribute, so I am taking the dot out
                        #at the source
                        if cell.value.endswith('.'):
                            goodValue = cell.value [:-1]
                        else:
                            goodValue = cell.value
                        print(goodValue)
                        dictionar_row['Matl Group'] = goodValue
                    elif self.cell_count == 6:
                        dictionar_row['Batch'] = cell.value
                    elif self.cell_count == 7:
                        dictionar_row['Quantity'] = cell.value
                    elif self.cell_count == 8:
                        dictionar_row['Unrestricted'] = cell.value
                    elif self.cell_count == 9:
                        dictionar_row['Expiration date'] = cell.value.date()
                    elif self.cell_count == 10:
                        dictionar_row['Description'] = cell.value
                self.table_as_list_of_dicts.append(dictionar_row)
        return self.table_as_list_of_dicts

# tester = AgedStock('StocksGB6X(20220415) (1).xlsx').runAll()
# print(tester)
